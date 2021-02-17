"""
Uso: PEP, pip y APIs
Creado: Andrés Hernández Mata
Version: 1.0
Python: 3.8.2
Fecha: 10 de Febrero del 2021
"""


from pyhunter import PyHunter
from openpyxl import Workbook
import getpass


def Busqueda(organizacion):
    resultado = hunter.domain_search (company = organizacion, limit = 10, emails_type = 'personal')
    return resultado


def GuardarInformacion(datosEncontrados,organizacion):
    libro = Workbook()
    hoja = libro.create_sheet(organizacion)
    libro.save("Hunter" + organizacion + ".xlsx")
    libro.active = 1
    excel = libro.active    
    excel.append(('Correo Electronico','Tipo','Nombre','Apellido','Telefono','Linkedin','Twitter'))
    count = 2
    emails = datosEncontrados['emails']    
    for x in emails:                                
        excel.cell(count,1,x['value'])
        excel.cell(count,2,x['type'])
        excel.cell(count,3,x['first_name'])
        excel.cell(count,4,x['last_name'])
        excel.cell(count,5,x['phone_number'])
        excel.cell(count,6,x['linkedin'])
        excel.cell(count,7,x['twitter']) 
        count+=1
    libro.save("Hunter" + organizacion + ".xlsx")
    

print("Script para buscar información")
apikey = getpass.getpass("Ingresa tu API key: ")
hunter = PyHunter(apikey)
orga = input("Dominio a investigar: ")
datosEncontrados = Busqueda(orga)
if datosEncontrados == None:
    exit()
else:
    print(datosEncontrados)
    print(type(datosEncontrados))
    GuardarInformacion(datosEncontrados,orga)